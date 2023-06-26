from openpyxl import load_workbook, workbook
import openpyxl
import pandas as pd
total_amt = []
momo_db_path = "MTN_Momo\momo.xlsx"

def create_acoount(phone_number,password,):
    data = pd.read_excel(momo_db_path).to_dict('records')
    for user in data:
        if user["PHONE"] == phone_number:
            return False

    workbook = load_workbook(momo_db_path)  
    sheet = workbook.active
    column_count = sheet.max_column
    row_count = sheet.max_row

    sheet[f"A{row_count+1}"] = phone_number
    sheet[f"B{row_count+1}"] = password
    
    workbook.save(momo_db_path)
    return True


def login_account(phone_number, password):
    data = pd.read_excel(momo_db_path).to_dict('records')
    for user in data:
        if user["PHONE"] == phone_number and user["PASSWORD"] == password:
            return user
    return False   


def enter_pin(password):
    data = pd.read_excel(momo_db_path).to_dict('records')
    for user in data:
        if user["PASSWORD"] == password:
            return user
    return False 


def register():
    print("======================")
    print("Registration Page")
    print("======================")
    phone_number = int(input("Phone_Number: "))
    password = int(input("PIN: "))
    user = create_acoount(phone_number, password)
    if user:
        login()
    else:
        print("Phone_Number OR Password Unavailable")
        print("Try Again")
        register()


def login():
    print("======================")
    print("Login Page")
    print("======================")   
    phone_number = int(input("Phone_Number: "))
    password = int(input("PIN: "))
    user = login_account(phone_number, password)
    if user:
        menu()
    else:
        print("Phone_Number OR Password Unavailable")
        print("Try Again")
        login()


def auth():
    print("1. Register 2. Login")

    reply = input("Enter: ")     
     
    if reply == "1":
        register()
    elif reply == "2":
        login()
    else:
        print("Wrong input")


def transfer_money():
    print("1. MoMo User")
    print("2. Other Networks")
    reply = input("Enter: ")  
   
    if reply == "1":
        print("")
        print("============")
        print("MoMo User")
        print("=============")
        trans = int(input("enter number:  "))
        amt = int(input("enter amount: "))
        ref = input("enter reference: ")
        if not total_amt:
            print("Your account is low")
            menu()
            exit()
        tot = int(total_amt[0]) - amt
        if amt > int(total_amt[0]):
            print("Your account is low")
            menu()
            exit()
        print("")
        print(f"Tranfer to - {trans} for GHS{amt} with Reference: {ref}, Total Amount is GHS{total_amt[0]}")
        password = int(input(" Entet MM PIN: "))
        user = enter_pin(password)
        if user:
            print(f"""Payment made for GHS{amt} to - {trans}. 
            Current Balance: GHS{tot}. Available Balance: GHS{tot}.Reference: {ref}.""")
        else:
            print("invalid option")
            menu()
    elif reply == "2":
        print("")
        print("=================")
        print("Other Networks")
        print("=================")
        print("1. Vodafone")
        print("2. AirtelTigo")
        ans = input("Enter: ")
        if ans == "1" or "2":
            print("")
            print("=================")
            trans = int(input("enter number:  "))
            amt = int(input("enter amount: "))
            ref = input("enter reference: ")
            if not total_amt:
                print("Your account is low")
                menu()
                exit()
            tot = int(total_amt[0]) - amt
            if amt > int(total_amt[0]):
                print("Your account is low")
                menu()
                exit()
            print("")
            print(f"Tranfer to - {trans} for GHS{amt} with Reference: {ref}, Total Amount is GHS{tot}")
            password = int(input(": "))
            user = enter_pin(password)
            if user:
                print(f"""Payment made for GHS{amt} to - 233{trans}. Current Balance: GHS{tot}.
                          Available Balance: GHS 100. Reference: {ref}.""")
        else:
            print("invalid option!!")
            menu()        
    else:
        print("invalid option")
        menu()


def check_balance():
    print("Fee is GHS 0.00. Enter MM PIN")
    password = int(input(": "))
    user = enter_pin(password)
    if user:
        if not total_amt:
            print("Your balance is GHS0")
            menu()
            print(f"Your balance is GHS{total_amt[0]}")
            menu()
    else:
        print("Invalid MM PIN code")    


def airtime_and_bundles():

    print("=================")
    print("Airtime & Bundles")
    print("==================")
    print("")
    print("1.Airtime")
    print("2.Bundles")
    reply = input("Enter: ")
    if reply == "1":
        print("")
        print("=========")
        print("Airtime")
        print("=========")
        print("1.  Self")
        print("2. Others")
        reply = input("Enter: ")
        if reply == "1":
            print("")
            print("============")
            print("Self")
            print("============")
            amnt = int(input("Enter Amount: "))
            if not total_amt:
                print("Your account is low")
                menu()
                exit()
            tot = int(total_amt[0]) - amnt
            if amnt > int(total_amt[0]):
                print("Your account is low")
                menu()
                exit()
            password = int(input("Enter MM PIN: "))
            user = enter_pin(password)
            if user:
                print(f"""Your payment of GHS{amnt} to MTN AIRTIME has been completed.Your new balance: GHS{tot}
                          Fee was GHS 0.00. REference: -.""")
            else:
                print("Invalid Option")
                menu()    
        elif reply == "2":
            trans = int(input("Enter number:  "))
            amnt = int(input("Enter amount: "))
            if not total_amt:
                print("Your account is low")
                menu()
                exit()
            tot = int(total_amt[0])- amnt
            if amnt > int(total_amt[0]):
                print("Your account is low")
                menu()
                exit()
            print("")
            print(f"""Send GHS{amnt} Airtime to {trans}. Fee is GHS 0.00.
            Enter MM PIN to confirm""")
            password = int(input(": "))
            user = enter_pin(password)
            if user:
                print(f"""Payment of GHS{amnt} of MTN AIRTIME made to {trans}. Current Balance: GHS{tot} """)
            else:
                print("Invalid Option")
                menu()   
        else:
            print("Invalid Option")
            menu()
    elif reply == "2":
        print("")
        print("=========")
        print("Internet Bundles")
        print("=========")
        print("1. Self")
        print("2. Others")
        reply = input("Enter: ")
        if reply == "1":
            print("")
            print("============")
            print("Self")
            print("============")
            amnt = int(input("Enter Amount: "))
            if not total_amt:
                print("Your account is low")
                menu()
                exit()
            if amnt > int(total_amt[0]):
                print("Your account is low")
                menu()
                exit() 
            total = amnt * 115.2
            tot = int(total_amt[0]) - amnt
            password = int(input("Enter MM PIN: "))
            user = enter_pin(password)
            if user:
                print(f"""Congratulations you have successfully purchased a {total}mb 
                data bundle @GHS{amnt}""")
            else:
                print("Invalid Option")
        elif reply == "2":
            trans = int(input("Enter number:  "))
            amnt = int(input("Enter amount: "))
            if not total_amt:
                print("Your account is low")
                menu()
                exit()
            if amnt > int(total_amt[0]):
                print("Your account is low")
                menu()
                exit()
            total = amnt * 115.2
            print("")
            print(f"""Send {total}mb data bundle to {trans}.Fee is GHS 0.00.Enter MM PIN to confirm""")
            password = int(input(": "))
            user = enter_pin(password)
            if user:
                print(f"""Congratulations you have successfully purchased a {total}mb 
                data bundle @GHS{amnt} for {trans}.""")
            else:
                print("Invalid Option")
                menu()          
    else:
        print("Invalid Option")
        menu()


def Change_and_reset_pin():
    print("")
    print("===================")
    print("Change & Reset PIN")
    print("==================")
    print("1. Change PIN")
    print("0. Back ")
    reply = input(": ")
    if reply == "1":
        print("Enter Old MM PIN code")
        password = int(input(": "))
        user = enter_pin(password)
        if user:
            print("")
            print("Enter New MM PIN code")
            new_code = input(": ")
            print("")
            print("Press 1 to Confirm your new MM PIN or Press 2 to Cancel")
            con = input(": ")
            if con == "1":
                user_db[0]["password"] = new_code
                print("")
                print("MM PIN code Updated")
                menu()
            elif con == "2":
                print("Canceled")
                menu() 
            else:
                print("Invalid Option")
                Change_and_reset_pin()
        else:
            print("Incorrect MM PIN code")
            menu()
    if reply == "0":
        menu()


def Deposite():
    print("")
    print("Enter amount")
    amt = int(input(": "))
    print("Enter MM PIN")
    password = int(input(": "))
    user = enter_pin(password)
    if user:
        total_amt.append(amt)
        print(f"You have successfully deposited GHS{amt} into your account")
        menu()
    else:
        print("Invalid MM PIN code")    



def menu():
    print(" ")
    print("      Menu")
    print("     ======")
    print("1. Transfer Money")
    print("2. Airtime & Bundle")
    print("3. Check Balance")
    print("4. Change & Reset PIN")
    print("5. Deposite")
    print("")
    Reply = input("Enter: ")
    if Reply == "3":
        check_balance()
    elif Reply == "1":
        transfer_money()
    elif Reply == "2":
        airtime_and_bundles()
    elif Reply == "4":
        Change_and_reset_pin()
    elif Reply == "5":
        Deposite()
            

def app():
    auth()   


app()